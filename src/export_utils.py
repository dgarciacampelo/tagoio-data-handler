from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Note: #00571B is a dark forest green that provides good contrast with white text
def generate_telemetry_excel(
    session: dict, telemetry_data: list[tuple], text_color: str = "FFFFFF", fill_color: str = "00571B"
) -> BytesIO:
    """Generates an XLSX file containing session metadata and tick-by-tick telemetry."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Strict type narrowing for the worksheet
    if not isinstance(ws, Worksheet):
        ws = wb.create_sheet()

    transaction_id = session["transaction_id"]
    ws.title = f"Audit_{transaction_id}"

    # Styling elements
    title_font = Font(bold=True, size=14, color=text_color)
    title_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(bold=True, color=text_color)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    meta_label_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # * 1. SUMMARY METADATA TABLE (Rows 1-5)
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Velo Energy: Charging Session Audit Report - Transaction {transaction_id}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center_align

    # Row 2: Station Data
    ws["A2"] = "Station:"
    ws["B2"] = f"{session['pool_code']} / {session['station_name']} [Conn: {session['connector_id']}]"
    ws["C2"] = "Start Date:"
    ws["D2"] = session["start_date"]
    ws["E2"] = "Charging Interval:"
    ws["F2"] = session["time_band"]

    # Row 3: Frozen Rates
    ws["A3"] = "Rates (€/kWh):"
    ws["B3"] = f"Valle: {session['rate_off_peak']:.4f} €"
    ws["C3"] = f"Llano: {session['rate_flat']:.4f} €"
    ws["D3"] = f"Punta: {session['rate_peak']:.4f} €"

    # Row 4: Final Aggregated Totals
    ws["A4"] = "Final Energy:"
    ws["B4"] = f"{session['total_energy_kwh']:.3f} kWh"
    ws["C4"] = "Final Cost:"
    ws["D4"] = f"{session['cost']:.2f} €"

    # Apply bold labels to metadata (using strict type checks)
    for r in range(2, 5):
        # Column 1 always contains a label (A2, A3, A4)
        cell_1 = ws.cell(row=r, column=1)
        if isinstance(cell_1, Cell):
            cell_1.font = meta_label_font

        # Column 3 contains a label ONLY for rows 2 and 4 (C2, C4)
        if r in [2, 4]:
            cell_3 = ws.cell(row=r, column=3)
            if isinstance(cell_3, Cell):
                cell_3.font = meta_label_font

        # Column 5 contains a label ONLY for row 2 (E2)
        if r == 2:
            cell_5 = ws.cell(row=r, column=5)
            if isinstance(cell_5, Cell):
                cell_5.font = meta_label_font

    # * 2. TELEMETRY TIME-SERIES TABLE (Row 7+)
    headers = [
        "Timestamp (UTC)",
        "Meter Value (Wh)",
        "Power (W)",
        "Reported Cost (€)",  # Cost reported by OCPP
        "Evaluated Cost (€)",  # Cost evaluated theoretically
        "Active Band",
        "Off-Peak Cum. (Wh)",
        "Flat Cum. (Wh)",
        "Peak Cum. (Wh)",
    ]

    header_row = 6
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)

        # Strict type narrowing for the header cells
        if isinstance(cell, Cell):
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        ws.column_dimensions[get_column_letter(col)].width = 20

    # Extract rates to simplify the calculation loop
    rates = {"off_peak": session["rate_off_peak"], "flat": session["rate_flat"], "peak": session["rate_peak"]}

    # Dictionary to translate the raw OCPP/Database bands to Spanish UI terms
    band_translation = {"Off-Peak": "Valle", "Flat": "Llano", "Peak": "Punta"}

    # Append data and calculate Evaluated Cost
    current_row = header_row + 1
    for row_data in telemetry_data:
        raw_timestamp = row_data[0]

        # Parse the ISO string to a naive Python datetime object
        try:
            # Replace 'Z' with +00:00 to ensure fromisoformat parses it safely
            clean_ts = raw_timestamp.replace("Z", "+00:00")
            dt_obj = datetime.fromisoformat(clean_ts)
            # Remove timezone info so Excel treats it cleanly without timezone warnings
            dt_val = dt_obj.replace(tzinfo=None)
        except Exception:  # Fallback to string if parsing fails
            dt_val = raw_timestamp

        meter_val = row_data[1]
        power = row_data[2]
        reported_cost = row_data[3]
        active_band = row_data[4]
        cum_off_peak = row_data[5]
        cum_flat = row_data[6]
        cum_peak = row_data[7]

        translated_band = band_translation.get(active_band, active_band)

        # Calculation: (Wh / 1000) * Rate
        evaluated_cost = (
            (cum_off_peak / 1000.0 * rates["off_peak"])
            + (cum_flat / 1000.0 * rates["flat"])
            + (cum_peak / 1000.0 * rates["peak"])
        )

        ws.append(
            [
                dt_val,
                meter_val,
                power,
                reported_cost,
                round(evaluated_cost, 4),
                f"{active_band}/{translated_band}",
                cum_off_peak,
                cum_flat,
                cum_peak,
            ]
        )  # Retain 4 decimal precision for debugging evaluated_cost

        # Apply specific datetime formatting to the Timestamp column
        date_cell = ws.cell(row=current_row, column=1)
        if isinstance(date_cell, Cell):
            date_cell.number_format = "yyyy-mm-dd hh:mm:ss"

        current_row += 1

    # * 3. EXCEL AUTO-FILTER
    # Enable filtering on the headers row (Row 6) down to the max row
    ws.auto_filter.ref = f"A{header_row}:I{ws.max_row}"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
